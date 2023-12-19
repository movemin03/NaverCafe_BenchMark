from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
import openpyxl
import os
from datetime import datetime
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import ttk, messagebox

root = tk.Tk()
root.title("네이버 카페 BenchMark")

keyword_label = ttk.Label(root, text="검색할 단어")
keyword_label.pack()
keyword_entry = ttk.Entry(root)
keyword_entry.pack()

member_min_label = ttk.Label(root, text="최저 멤버 수(선택사항: 기본값 0 )")
member_min_label.pack()
member_min_entry = ttk.Entry(root)
member_min_entry.insert(0, "0")
member_min_entry.pack()

exception_label = ttk.Label(root, text="예외처리할 키워드(선택사항: 콤마(,)로 구분)")
exception_label.pack()
exception_entry = ttk.Entry(root)
exception_entry.pack()

radio_var = tk.IntVar()
radio1 = ttk.Radiobutton(root, text='랭킹순', variable=radio_var, value='0')
radio2 = ttk.Radiobutton(root, text='정확도순', variable=radio_var, value='1')
radio1.pack()
radio2.pack()


Result_Viewlabel = ttk.LabelFrame(text="실행내용")
Result_Viewlabel.pack()

Result_Viewlabel_Scrollbar = tk.Listbox(Result_Viewlabel, selectmode="extended", width=80, height=20, font=('Normal', 9), yscrollcommand=tk.Scrollbar(Result_Viewlabel).set)
Result_Viewlabel_Scrollbar.pack(side=tk.LEFT, fill=tk.BOTH)

pb_type = tk.DoubleVar()
progress_bar = ttk.Progressbar(orient="horizontal", length=400, mode="determinate", maximum=100,variable=pb_type)

progress = 0
execute_num = 0

def work():
    global keyword
    keyword = keyword_entry.get()
    if keyword == "":
        messagebox.showinfo(title="알림", message="검색어가 입력되지 않았습니다")
    else:
        global member, member_min, exception_keyword, start_pg
        progress_bar.pack()
        member_min = int(member_min_entry.get())
        exception_keyword_input = exception_entry.get().replace(" ", "")
        exception_keyword = exception_keyword_input.split(',')
        start_pg = 1
        order = radio_var.get()

        if order == "1":
            url = "https://section.cafe.naver.com/ca-fe/home/search/cafes?q=" + keyword + "&t=1701084772896"
            Result_Viewlabel_Scrollbar.insert(tk.END, "정확도순으로 검색합니다")
        else:
            url = "https://section.cafe.naver.com/ca-fe/home/search/cafes?q=" + keyword + "&t=1700821664452&od=1"
            Result_Viewlabel_Scrollbar.insert(tk.END, "랭킹순으로 검색합니다")

        Result_Viewlabel_Scrollbar.see(tk.END)
        progress = 30
        pb_type.set(progress)
        progress_bar.update()

        # 데이터프레임 생성
        global data
        data = {
            'Link': [],
            'Cafe Title': [],
            'Member Num': [],
            'Ranking': [],
            'NewPostNum': [],
            'TotalPostNum': [],
            'TotalScore': []
        }

        start_time = datetime.now()

        # 크롬드라이버 실행
        Result_Viewlabel_Scrollbar.insert(tk.END, "크롬 로딩 중입니다... 컴퓨터 사양에 따라 실행에 시간이 걸릴 수 있습니다")
        progress = 50
        pb_type.set(progress)
        progress_bar.update()

        #root.update_idletasks() 화면 전체
        Result_Viewlabel_Scrollbar.update()
        options = webdriver.ChromeOptions()
        options.add_argument("headless")
        # driver = webdriver.Chrome() 크롬창 headless 모드를 해제하려면 이것 사용
        global driver
        driver = webdriver.Chrome(options=options)
        driver.get(url)

        global wait
        wait = WebDriverWait(driver, 10)
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="mainContainer"]/div[2]/div[1]/div[2]/div[1]/div[1]')))
        end_time = datetime.now()
        execution_time = end_time - start_time
        Result_Viewlabel_Scrollbar.insert(tk.END, f"크롬 로딩 시간: {execution_time}")
        progress = 70
        pb_type.set(progress)
        progress_bar.update()

        try:
            cafe_total = int(driver.find_element(By.XPATH, '//*[@id="mainContainer"]/div[2]/div[1]/div[2]/div[1]/div[1]').text.replace(',', ''))
        except ValueError:
            try:
                int(driver.find_element(By.XPATH,'//*[@id="mainContainer"]/div[2]/div[1]/div[2]/div[1]/div[1]').text.replace(',',''))
            except ValueError:
                Result_Viewlabel_Scrollbar.insert(tk.END, "알 수 없는 오류로 페이지 로딩에 실패했습니다. 검색 버튼을 다시 눌러주세요")
        try:
            max_pg = round(cafe_total / 12)
        except UnboundLocalError:
            "알 수 없는 오류로 페이지 로딩에 실패했습니다. 검색 버튼을 다시 눌러주세요"
        Result_Viewlabel_Scrollbar.insert(tk.END, "최대 탐색가능한 페이지는 '" + str(max_pg) + "'입니다")
        Result_Viewlabel_Scrollbar.insert(tk.END, "어디까지 탐색할 것인지 값을 넣어주세요")
        Result_Viewlabel_Scrollbar.insert(tk.END, "다시 검색하시려면, 검색어 변경 후 검색 버튼을 다시 눌러주세요")
        Result_Viewlabel_Scrollbar.insert(tk.END, "계속 하려면, 마지막 페이지 입력하고 계속 진행 버튼 클릭")
        Result_Viewlabel_Scrollbar.see(tk.END)
        progress = 100
        pb_type.set(progress)
        progress_bar.update()

        messagebox.showinfo(title="알림", message="최대 탐색가능한 페이지는 '" + str(max_pg) + "'입니다. 탐색할 마지막 페이지를 입력하고 계속진행 버튼 클릭")

        global execute_num
        if execute_num < 1:
            last_pg_label = ttk.Label(root, text="마지막 페이지")
            last_pg_label.pack()
            global last_pg_entry
            last_pg_entry = ttk.Entry(root)
            last_pg_entry.pack()
            button_2 = ttk.Button(root, text="계속 진행", command=search_loop)
            button_2.pack()
            execute_num = execute_num + 1
        else:
            pass

        global button
        root.update_idletasks()


button = ttk.Button(root, text="검색", command=work)
button.pack()


def search(search_pg):
    Result_Viewlabel_Scrollbar.insert(tk.END, "탐색 시작")
    Result_Viewlabel_Scrollbar.update()
    for item_num in range(1, 13):
        item_num_link = "//div[@class='item_list']//div[@class='CafeItem'][" + str(item_num) + "]//a"
        Result_Viewlabel_Scrollbar.insert(tk.END, str(search_pg) + "페이지 중 " + str(item_num) + "/12항목 = 전체 중 " + str(12*(search_pg-1)+item_num) + "번 검색 중")
        Result_Viewlabel_Scrollbar.update()
        Result_Viewlabel_Scrollbar.see(tk.END)
        progress = (12 * (int(search_pg) - 1) + int(item_num)) / (12 * int(last_pg)) * 100
        pb_type.set(progress)
        progress_bar.update()
        try:
            wait.until(EC.presence_of_element_located((By.XPATH, item_num_link)))
            link = driver.find_element(By.XPATH, item_num_link).get_attribute("href")
            cafe_title = driver.find_element(By.XPATH, item_num_link + "/div[@class='cafe_info_area']/strong[@class='cafe_name']").text
            member_num = int(driver.find_element(By.XPATH, item_num_link + "/div[@class='cafe_info_area']/div[@class='cafe_info_detail']/*[2]/*[2]").text.replace(",", ""))
            ranking = driver.find_element(By.XPATH, item_num_link + "/div[@class='cafe_info_area']/div[@class='cafe_info_detail']/*[3]/*[2]").text
            ranking_main = ranking[:2]
            ranking_dic = {'숲': 6.83, '나무': 6, '열매': 5, '가지': 4, '잎새': 3, '새싹': 2, '씨앗': 1}
            if ranking_main in ranking_dic:
                ranking_main_num = ranking_dic[ranking_main]
            ranking_sub = int(ranking[2:3])
            if ranking_sub == "":
                ranking_sub = 1
            post_num = driver.find_element(By.XPATH, item_num_link + "/div[@class='cafe_info_area']/div[@class='cafe_info_detail']/*[4]/*[2]").text
            new_post_num, total_post_num = post_num.split("/")
            new_post_num = int(new_post_num.replace(",", "").replace(" ", ""))
            total_post_num = int(total_post_num.replace(",", "").replace(" ", ""))

            # 점수계산 = (멤버 수 * 0.3) + ((랭킹대분류+1/6*(랭킹소분류-1)) * 0.6 * 100) + (새로운 글 수 * 0.015) + (전체 글 수 * 0.085)
            score = round((member_num * 0.3) + ((ranking_main_num + 1/6 * (ranking_sub - 1))*0.6*100) + (new_post_num * 0.015) + (total_post_num * 0.085))

            if any(word in cafe_title for word in exception_keyword) and exception_keyword != ['']:
                Result_Viewlabel_Scrollbar.insert(tk.END, "예외 처리: Cafe Title 에 예외처리한 단어가 포함되어 있으므로 엑셀 파일에는 포함되지 않습니다")
                Result_Viewlabel_Scrollbar.update()
            elif member_min != 0 and member_num < member_min:
                Result_Viewlabel_Scrollbar.insert(tk.END, "예외 처리: 회원 수가 사용자가 지정한 최저 사용자 수인 " + str(member_min) + "보다 적은 " + str(member_num) + "명이라 엑셀 파일에 포함하지 않습니다")
                Result_Viewlabel_Scrollbar.update()

            else:
                data['Link'].append(link)
                data['Cafe Title'].append(cafe_title)
                data['Member Num'].append(str(member_num))
                data['Ranking'].append(ranking)
                data['NewPostNum'].append(str(new_post_num))
                data['TotalPostNum'].append(str(total_post_num))
                data['TotalScore'].append(str(score))
                Result_Viewlabel_Scrollbar.insert(tk.END, "검색 완료")
                Result_Viewlabel_Scrollbar.update()
        except:
            Result_Viewlabel_Scrollbar.insert(tk.END, "조회된 항목이 없습니다")
            Result_Viewlabel_Scrollbar.update()
            pass


def search_loop():
    start_time = datetime.now()
    global last_pg
    try:
        last_pg = int(last_pg_entry.get())
    except ValueError:
        messagebox.showinfo(title="알림", message="마지막 페이지가 입력되지 않았습니다")

    progress = 0
    pb_type.set(progress)
    progress_bar.update()
    Result_Viewlabel_Scrollbar.insert(tk.END, "시작페이지는 " + str(start_pg) + ", 마지막페이지는 " + str(last_pg) + "로 지정되었습니다")
    for search_pg in range(start_pg, last_pg+1):
        wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='item_list']//div[@class='CafeItem'][1]//a")))
        if search_pg % 10 == 1 and search_pg != 1:
            Result_Viewlabel_Scrollbar.insert(tk.END, "페이지 넘어갑니다")
            driver.find_element(By.XPATH, '//button[@type="button" and contains(@class, "btn type_next")]').click()
            wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='item_list']//div[@class='CafeItem'][1]//a")))
        pg_button = driver.find_element(By.XPATH, f"//div[@class='ArticlePaginate']//button[contains(text(), '{search_pg}')]")
        pg_button.click()
        wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class='item_list']//div[@class='CafeItem'][1]//a")))
        Result_Viewlabel_Scrollbar.insert(tk.END, "\n페이지" + str(search_pg) + "을(를) 탐색합니다\n")
        search(search_pg)

    # 데이터프레임 생성
    df = pd.DataFrame(data)
    df['TotalScore'] = pd.to_numeric(df['TotalScore'])
    median = df['TotalScore'].drop_duplicates().median()

    file_time = datetime.today().strftime("%Y%m%d_%H%M")
    user = os.getlogin()
    file_path = 'C:\\Users\\' + user + '\\Desktop\\cafe_data_' + keyword + "_" + file_time
    sorted_df = df.sort_values(by='TotalScore', ascending=False)

    # 엑셀 파일로 저장
    with pd.ExcelWriter(file_path + ".xlsx") as writer:
        sorted_df.to_excel(writer, index=False, sheet_name='Sheet1')

        # 중간값 셀을 빨간색으로 표시
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        median_cell = worksheet.cell(row=1, column=sorted_df.columns.get_loc('TotalScore') + 2)
        median_cell.fill = red_fill
        median_cell.value = median

    Result_Viewlabel_Scrollbar.insert(tk.END, "\n" + file_path + ".xlsx 에 저장 완료되었습니다")
    end_time = datetime.now()
    execution_time = end_time - start_time
    Result_Viewlabel_Scrollbar.insert(tk.END, f"검색 소요 시간: {execution_time}")
    Result_Viewlabel_Scrollbar.insert(tk.END, "total_Score 의 중복 제외 중간값은 " + str(median) + " 입니다")
    Result_Viewlabel_Scrollbar.see(tk.END)

    pb_type.set(100)
    progress_bar.update()
    messagebox.showinfo(title="알림", message="검색이 완료되어 " + file_path + ".xlsx 에 저장 완료되었습니다")

root.mainloop()
